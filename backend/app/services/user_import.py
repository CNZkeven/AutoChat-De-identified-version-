from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

IMPORT_HEADERS = ["学号", "姓名", "专业", "年级", "性别"]
ALLOWED_GENDERS = {"男", "女", "未知"}


def build_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "用户导入"
    ws.append(IMPORT_HEADERS)
    ws.append(["20230001", "张三", "计算机科学与技术", 2023, "男"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_import_file(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active
    errors: list[str] = []
    rows: list[dict[str, Any]] = []

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if headers[: len(IMPORT_HEADERS)] != IMPORT_HEADERS:
        errors.append("导入模板表头不匹配，请使用系统模板")
        return [], errors

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in row[: len(IMPORT_HEADERS)]]
        record = dict(zip(IMPORT_HEADERS, values))
        student_no = str(record.get("学号") or "").strip()
        full_name = str(record.get("姓名") or "").strip()
        major = str(record.get("专业") or "").strip()
        grade_raw = record.get("年级")
        gender = str(record.get("性别") or "").strip() or "未知"

        if not student_no:
            errors.append(f"第 {idx} 行：学号为空")
            continue
        if not full_name:
            errors.append(f"第 {idx} 行：姓名为空")
            continue
        if not major:
            errors.append(f"第 {idx} 行：专业为空")
            continue
        if grade_raw is None or str(grade_raw).strip() == "":
            errors.append(f"第 {idx} 行：年级为空")
            continue
        try:
            grade = int(str(grade_raw).strip())
        except ValueError:
            errors.append(f"第 {idx} 行：年级不是数字")
            continue
        if gender not in ALLOWED_GENDERS:
            errors.append(f"第 {idx} 行：性别需为 男/女/未知")
            continue

        rows.append(
            {
                "username": student_no,
                "full_name": full_name,
                "major": major,
                "grade": grade,
                "gender": gender,
            }
        )

    return rows, errors
